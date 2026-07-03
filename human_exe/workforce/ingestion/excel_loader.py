"""Excel loader with optional openpyxl dependency."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from human_exe.workforce.ingestion.dependency_check import ensure_openpyxl_available


def load_excel(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    ensure_openpyxl_available()
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        result.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
    return result
